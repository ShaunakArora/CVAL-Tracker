from flask import Flask, render_template, jsonify, send_from_directory, request, redirect, url_for, flash, session, Response
from dotenv import load_dotenv
load_dotenv()
import os
import json
import io
import re
import secrets
import logging
import click
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from flask_wtf.csrf import CSRFProtect
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, extract, text
import dash
from dash import dcc, html
import plotly.express as px
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))

# Security Configuration
secret_key = os.environ.get('SECRET_KEY')
if not secret_key:
    if os.environ.get('FLASK_ENV') == 'production':
        raise RuntimeError("SECRET_KEY environment variable is required in production.")
    else:
        import secrets
        secret_key = secrets.token_hex(24)
app.secret_key = secret_key
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'

# Database Configuration
db_uri = os.environ.get('DATABASE_URL')
if db_uri and db_uri.startswith("postgres://"):
    db_uri = db_uri.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_uri or f"sqlite:///{os.path.join(BASE_DIR, 'app.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 280,
}
db = SQLAlchemy(app)

csrf = CSRFProtect(app)

# Configure permanent session lifetime for Session Expiry Policy
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Database schema upgrade at startup: adds lockout columns to existing schema (moved below model definitions)

# --- Security Helper Functions ---

# In-memory Rate Limiting for Login endpoint
# SCALABILITY NOTE (Issue #7): This dict is in-memory, per-process.
# With gunicorn -w 1 (current Procfile) this works correctly.
# If worker count is increased, each worker gets its own copy of this dict,
# making the effective limit 10 * num_workers instead of 10.
# Before scaling workers, move rate-limit state to Redis or the database.
login_attempts = {}  # ip_address -> list of timestamps

def is_rate_limited(ip_address, limit=10, period=60):
    """Checks if an IP address has exceeded the rate limit (limit requests per period seconds)."""
    now = datetime.now()
    timestamps = login_attempts.get(ip_address, [])
    # Filter timestamps within the period
    timestamps = [t for t in timestamps if (now - t).total_seconds() < period]
    login_attempts[ip_address] = timestamps
    if len(timestamps) >= limit:
        return True
    return False

def record_attempt(ip_address):
    now = datetime.now()
    if ip_address not in login_attempts:
        login_attempts[ip_address] = []
    login_attempts[ip_address].append(now)

# Password Complexity Policy validator
def validate_password_complexity(password):
    """Enforces minimum length of 8, at least one uppercase, lowercase, digit, and special char."""
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."
    if not re.search(r"[A-Z]", password):
        return False, "Password must contain at least one uppercase letter."
    if not re.search(r"[a-z]", password):
        return False, "Password must contain at least one lowercase letter."
    if not re.search(r"\d", password):
        return False, "Password must contain at least one digit."
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        return False, "Password must contain at least one special character."
    return True, ""

# --- Security Global Hooks & Error Handlers ---

@app.before_request
def enforce_https():
    """Redirects HTTP requests to HTTPS in production environment."""
    if os.environ.get('FLASK_ENV') == 'production':
        proto = request.headers.get('X-Forwarded-Proto', 'http')
        if proto != 'https':
            url = request.url.replace('http://', 'https://', 1)
            return redirect(url, code=301)

@app.after_request
def add_security_headers(response):
    """Enforces secure browser settings via HTTP headers."""
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    if os.environ.get('FLASK_ENV') == 'production' or request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # Content Security Policy configured to allow bootstrap and standard CDNs used by this dashboard
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://stackpath.bootstrapcdn.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://stackpath.bootstrapcdn.com https://cdnjs.cloudflare.com; "
        "font-src 'self' https://cdnjs.cloudflare.com; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    return response

@app.errorhandler(500)
def internal_server_error(e):
    """Generic secure error envelope to avoid exposing stack trace details to users."""
    app.logger.error(f"Internal Server Error (500): {e}", exc_info=True)
    try:
        add_system_alert("Internal system error occurred (500).")
    except Exception:
        pass
    if request.path.startswith('/api/') or request.headers.get('Content-Type') == 'application/json':
        return jsonify({"error": "An internal server error occurred."}), 500
    return render_template('error.html', error_message="An internal server error occurred. Please try again later or contact support."), 500

# --- SQLAlchemy Models ---

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(50), unique=True, nullable=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='employee')
    department = db.Column(db.String(80))
    shift = db.Column(db.String(50))
    location = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # Security tracking fields for lockout policy
    failed_login_attempts = db.Column(db.Integer, default=0, nullable=False)
    lockout_until = db.Column(db.DateTime, nullable=True)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Log(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_member = db.Column(db.String(80), nullable=False, index=True)
    function = db.Column(db.String(100))
    date = db.Column(db.Date, index=True)   
    file_number = db.Column(db.String(100))
    status = db.Column(db.String(100))
    tier1_escalation_reason = db.Column(db.String(200))
    im_escalation_reason = db.Column(db.String(200))
    department = db.Column(db.String(80))
    comments = db.Column(db.Text)
    # Issue #3 fix: Integer instead of String so SUM/aggregations work correctly.
    count = db.Column(db.Integer)
    bucket = db.Column(db.String(100))
    time = db.Column(db.String(50))
    production_task = db.Column(db.String(100))
    month = db.Column(db.String(50))

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Alert(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    message = db.Column(db.String(500), nullable=False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Function(db.Model):
    __tablename__ = 'functions'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Department(db.Model):
    __tablename__ = 'department'
    id = db.Column(db.Integer, primary_key=True)
    dept_name = db.Column(db.String(100), unique=True, nullable=False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)


with app.app_context():
    try:
        db.create_all()

        # --- Issue #1: Warn if .env file is present in production ---
        if os.environ.get('FLASK_ENV') == 'production':
            env_path = os.path.join(BASE_DIR, '.env')
            if os.path.exists(env_path):
                logging.critical(
                    "SECURITY WARNING: .env file detected inside the deployment folder (%s). "
                    "This file contains live credentials (DATABASE_URL, SECRET_KEY). "
                    "Remove it immediately and store secrets via platform environment variables.",
                    env_path
                )

        # Add columns manually in case the table already exists but doesn't have them
        try:
            db.session.execute(text('ALTER TABLE "user" ADD COLUMN failed_login_attempts INTEGER DEFAULT 0;'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        try:
            db.session.execute(text('ALTER TABLE "user" ADD COLUMN lockout_until TIMESTAMP;'))
            db.session.commit()
        except Exception:
            db.session.rollback()

        # --- Issue #3: Migrate count column from String to Integer (Postgres only) ---
        try:
            db.session.execute(text(
                'ALTER TABLE log ALTER COLUMN count TYPE INTEGER USING '
                'CASE WHEN count ~ \'^\'\'^[0-9]+$\'\' THEN count::INTEGER ELSE 1 END;'
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()  # Silently ignore on SQLite or if already Integer

        # --- Issue #4: Partial unique index to prevent duplicate 'In Progress' rows ---
        # This creates a DB-level constraint so even concurrent requests cannot create
        # two simultaneous 'In Progress' logs for the same team_member.
        try:
            db.session.execute(text(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_log_team_member_in_progress "
                "ON log (team_member) WHERE status = 'In Progress';"
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()  # Silently ignore on DBs that don't support partial indexes

        # Synchronize departments from logs to Department table
        try:
            log_depts = db.session.query(Log.department).distinct().all()
            log_depts = [d[0] for d in log_depts if d[0] and d[0].strip()]
            existing_depts = {d.dept_name for d in Department.query.all()}
            added = False
            for dept_name in log_depts:
                if dept_name not in existing_depts:
                    db.session.add(Department(dept_name=dept_name))
                    added = True
            if added:
                db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Error synchronizing departments at startup: {e}")
    except Exception as e:
        print(f"Error checking/updating database schema: {e}")


def add_system_alert(message):
    """Adds a system alert to the database and keeps the latest 50."""
    new_alert = Alert(message=message, timestamp=datetime.now())
    db.session.add(new_alert)

    # Keep only the last 50 alerts
    try:
        alert_count = Alert.query.count()
        if alert_count > 50:
            oldest_alerts = Alert.query.order_by(Alert.timestamp.asc()).limit(alert_count - 50).all()
            for alert in oldest_alerts:
                db.session.delete(alert)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error managing alerts: {e}")

def login_required(f):
    """Decorator to ensure a user is logged in."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('You need to be logged in to view this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to ensure a user is an admin."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('You need to be logged in to view this page.', 'warning')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('You do not have permission to access this page.', 'danger')
            # Redirect non-admins to their own dashboard
            return redirect(url_for('employee_dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def init_dashboard(app):
    """Create a Plotly Dash dashboard."""
    dash_app = dash.Dash(
        server=app,
        routes_pathname_prefix="/admin/analytics/",
        external_stylesheets=[
            "https://stackpath.bootstrapcdn.com/bootstrap/4.3.1/css/bootstrap.min.css"
        ],
        suppress_callback_exceptions=True
    )

    # Protect Dash views
    for view_func_name, view_func in app.view_functions.items():
        if view_func_name.startswith(dash_app.config["routes_pathname_prefix"]):
            app.view_functions[view_func_name] = admin_required(view_func)

    # Create Dash layout
    def create_layout():
        # Issue #6: Use SQL aggregations instead of loading all rows into Python.
        # Check if any data exists at all
        total_logs = db.session.query(func.count(Log.id)).scalar() or 0
        if total_logs == 0:
            return html.Div([
                html.H1("Analytics Dashboard"),
                html.P("No data available to display.")
            ], className="container")

        # 1. KPI: completion rate via SQL
        completed_logs = db.session.query(func.count(Log.id)).filter(
            Log.status.in_(['Completed', 'Approved'])
        ).scalar() or 0
        completion_rate = (completed_logs / total_logs) * 100 if total_logs > 0 else 0

        # Top employee by log count (SQL)
        top_emp_row = db.session.query(
            Log.team_member, func.count(Log.id).label('cnt')
        ).group_by(Log.team_member).order_by(func.count(Log.id).desc()).first()
        top_employee = top_emp_row.team_member if top_emp_row else "N/A"

        # 2. Time series: logs per date (SQL GROUP BY)
        time_rows = db.session.query(
            Log.date, func.count(Log.id).label('cnt')
        ).filter(Log.date.isnot(None)).group_by(Log.date).order_by(Log.date).all()
        logs_over_time = pd.DataFrame(
            [{'date': r.date, 'count': r.cnt} for r in time_rows]
        )
        if logs_over_time.empty:
            return html.Div([
                html.H1("Analytics Dashboard"),
                html.P("No data with valid dates available.")
            ], className="container")
        time_series_fig = px.line(
            logs_over_time, x='date', y='count',
            title='Total Logs Over Time', labels={'date': 'Date', 'count': 'Number of Logs'}
        )

        # 3. Top 10 functions (SQL)
        func_rows = db.session.query(
            Log.function, func.count(Log.id).label('cnt')
        ).filter(Log.function.isnot(None)).group_by(Log.function
        ).order_by(func.count(Log.id).desc()).limit(10).all()
        func_df = pd.DataFrame([{'function': r.function, 'cnt': r.cnt} for r in func_rows])
        func_df = func_df.sort_values('cnt', ascending=True)
        top_functions_fig = px.bar(
            func_df, x='cnt', y='function', orientation='h',
            title='Top 10 Functions', labels={'cnt': 'Count', 'function': 'Function'}
        )

        # Top 10 employees (SQL)
        emp_rows = db.session.query(
            Log.team_member, func.count(Log.id).label('cnt')
        ).filter(Log.team_member.isnot(None)).group_by(Log.team_member
        ).order_by(func.count(Log.id).desc()).limit(10).all()
        emp_df = pd.DataFrame([{'team_member': r.team_member, 'cnt': r.cnt} for r in emp_rows])
        emp_df = emp_df.sort_values('cnt', ascending=True)
        top_employees_fig = px.bar(
            emp_df, x='cnt', y='team_member', orientation='h',
            title='Top 10 Employees by Logs', labels={'cnt': 'Count', 'team_member': 'Employee'}
        )

        # 4. Function distribution donut (SQL)
        all_func_rows = db.session.query(
            Log.function, func.count(Log.id).label('cnt')
        ).filter(Log.function.isnot(None)).group_by(Log.function).all()
        dist_df = pd.DataFrame([{'function': r.function, 'cnt': r.cnt} for r in all_func_rows])
        function_dist_fig = px.pie(
            dist_df, values='cnt', names='function',
            title='Functions Distribution', hole=0.4
        )

        layout = html.Div(className="container-fluid", children=[
            html.H1("Analytics Dashboard", className="my-4"),

            # KPI Cards
            html.Div(className="row", children=[
                html.Div(className="col-md-4", children=[
                    html.Div(className="card text-white bg-primary mb-3", children=[
                        html.Div(className="card-header", children="Total Logs"),
                        html.Div(className="card-body", children=[html.H4(f"{total_logs}", className="card-title")])
                    ])
                ]),
                html.Div(className="col-md-4", children=[
                    html.Div(className="card text-white bg-success mb-3", children=[
                        html.Div(className="card-header", children="Completion Rate"),
                        html.Div(className="card-body", children=[html.H4(f"{completion_rate:.2f}%", className="card-title")])
                    ])
                ]),
                html.Div(className="col-md-4", children=[
                    html.Div(className="card text-white bg-info mb-3", children=[
                        html.Div(className="card-header", children="Top Employee (by logs)"),
                        html.Div(className="card-body", children=[html.H4(top_employee, className="card-title")])
                    ])
                ]),
            ]),

            # Time Series
            html.Div(className="row", children=[
                html.Div(className="col", children=[
                    dcc.Graph(figure=time_series_fig)
                ])
            ]),

            # Bar Charts
            html.Div(className="row mt-4", children=[
                html.Div(className="col-md-6", children=[
                    dcc.Graph(figure=top_functions_fig)
                ]),
                html.Div(className="col-md-6", children=[
                    dcc.Graph(figure=top_employees_fig)
                ])
            ]),

            # Donut Chart
            html.Div(className="row mt-4", children=[
                html.Div(className="col-md-8 offset-md-2", children=[
                    dcc.Graph(figure=function_dist_fig)
                ])
            ])
        ])
        return layout

    dash_app.layout = create_layout

    return dash_app.server

def init_daily_dashboard(app):
    """Create a Plotly Dash dashboard for Daily Analytics."""
    dash_app = dash.Dash(
        server=app,
        routes_pathname_prefix="/admin/daily_analytics/",
        external_stylesheets=[
            "https://stackpath.bootstrapcdn.com/bootstrap/4.3.1/css/bootstrap.min.css"
        ],
        suppress_callback_exceptions=True
    )

    # Protect Dash views
    for view_func_name, view_func in app.view_functions.items():
        if view_func_name.startswith(dash_app.config["routes_pathname_prefix"]):
            app.view_functions[view_func_name] = admin_required(view_func)

    # Create Dash layout
    def create_layout():
        # Issue #6: Use SQL aggregations instead of loading all rows for the day.
        today = datetime.now().date()
        total_logs = db.session.query(func.count(Log.id)).filter(Log.date == today).scalar() or 0

        if total_logs == 0:
            return html.Div([
                html.H1(f"Daily Analytics ({today})"),
                html.P("No data available for today.")
            ], className="container")

        # KPI: completion rate (SQL)
        completed_logs = db.session.query(func.count(Log.id)).filter(
            Log.date == today,
            Log.status.in_(['Completed', 'Approved'])
        ).scalar() or 0
        completion_rate = (completed_logs / total_logs) * 100 if total_logs > 0 else 0

        # Top employee today (SQL)
        top_emp_row = db.session.query(
            Log.team_member, func.count(Log.id).label('cnt')
        ).filter(Log.date == today
        ).group_by(Log.team_member).order_by(func.count(Log.id).desc()).first()
        top_employee = top_emp_row.team_member if top_emp_row else "N/A"

        # Top 10 functions today (SQL)
        func_rows = db.session.query(
            Log.function, func.count(Log.id).label('cnt')
        ).filter(Log.date == today, Log.function.isnot(None)
        ).group_by(Log.function).order_by(func.count(Log.id).desc()).limit(10).all()
        func_df = pd.DataFrame([{'function': r.function, 'cnt': r.cnt} for r in func_rows])
        func_df = func_df.sort_values('cnt', ascending=True)
        top_functions_fig = px.bar(
            func_df, x='cnt', y='function', orientation='h',
            title='Top 10 Functions Today', labels={'cnt': 'Count', 'function': 'Function'}
        )

        # Top 10 employees today (SQL)
        emp_rows = db.session.query(
            Log.team_member, func.count(Log.id).label('cnt')
        ).filter(Log.date == today, Log.team_member.isnot(None)
        ).group_by(Log.team_member).order_by(func.count(Log.id).desc()).limit(10).all()
        emp_df = pd.DataFrame([{'team_member': r.team_member, 'cnt': r.cnt} for r in emp_rows])
        emp_df = emp_df.sort_values('cnt', ascending=True)
        top_employees_fig = px.bar(
            emp_df, x='cnt', y='team_member', orientation='h',
            title='Top 10 Employees Today', labels={'cnt': 'Count', 'team_member': 'Employee'}
        )

        # Function distribution today (SQL)
        all_func_rows = db.session.query(
            Log.function, func.count(Log.id).label('cnt')
        ).filter(Log.date == today, Log.function.isnot(None)).group_by(Log.function).all()
        dist_df = pd.DataFrame([{'function': r.function, 'cnt': r.cnt} for r in all_func_rows])
        function_dist_fig = px.pie(
            dist_df, values='cnt', names='function',
            title='Functions Distribution Today', hole=0.4
        )

        layout = html.Div(className="container-fluid", children=[
            html.Div(className="row align-items-center my-4", children=[
                html.Div(className="col-md-9", children=[
                    html.H1(f"Daily Analytics ({today})"),
                ]),
                html.Div(className="col-md-3 text-md-end", children=[
                    html.A(
                        [html.I(className="fas fa-file-excel me-2"), "Export to Excel"],
                        href="/admin/daily_analytics/export",
                        className="btn btn-success",
                    )
                ])
            ]),

            # KPI Cards
            html.Div(className="row", children=[
                html.Div(className="col-md-4", children=[
                    html.Div(className="card text-white bg-primary mb-3", children=[
                        html.Div(className="card-header", children="Total Logs Today"),
                        html.Div(className="card-body", children=[html.H4(f"{total_logs}", className="card-title")])
                    ])
                ]),
                html.Div(className="col-md-4", children=[
                    html.Div(className="card text-white bg-success mb-3", children=[
                        html.Div(className="card-header", children="Completion Rate"),
                        html.Div(className="card-body", children=[html.H4(f"{completion_rate:.2f}%", className="card-title")])
                    ])
                ]),
                html.Div(className="col-md-4", children=[
                    html.Div(className="card text-white bg-info mb-3", children=[
                        html.Div(className="card-header", children="Top Employee Today"),
                        html.Div(className="card-body", children=[html.H4(top_employee, className="card-title")])
                    ])
                ]),
            ]),

            # Bar Charts
            html.Div(className="row mt-4", children=[
                html.Div(className="col-md-6", children=[
                    dcc.Graph(figure=top_functions_fig)
                ]),
                html.Div(className="col-md-6", children=[
                    dcc.Graph(figure=top_employees_fig)
                ])
            ]),

            # Donut Chart
            html.Div(className="row mt-4", children=[
                html.Div(className="col-md-8 offset-md-2", children=[
                    dcc.Graph(figure=function_dist_fig)
                ])
            ])
        ])
        return layout

    dash_app.layout = create_layout

    return dash_app.server

@app.route('/admin/daily_analytics/export')
@admin_required
def export_daily_analytics():
    """Exports the daily analytics dashboard to a formatted Excel file."""
    today = datetime.now().date()
    logs = Log.query.filter(Log.date == today).all()

    if not logs:
        flash('No data available for today to export.', 'warning')
        return redirect('/admin/daily_analytics/')

    df = pd.DataFrame([
        {"team_member": log.team_member, "function": log.function, "date": log.date, "status": log.status}
        for log in logs
    ])

    if df.empty:
        flash('No data available to display.', 'warning')
        return redirect('/admin/daily_analytics/')

    # --- 1. Calculate KPIs ---
    total_logs = len(df)
    completed_logs = df[df['status'].isin(['Completed', 'Approved'])].shape[0]
    completion_rate = (completed_logs / total_logs) * 100 if total_logs > 0 else 0
    top_employee_series = df['team_member'].mode()
    top_employee = top_employee_series[0] if not top_employee_series.empty else "N/A"

    # --- 2. Generate Figures ---
    top_functions = df['function'].value_counts().nlargest(10).sort_values(ascending=True)
    top_functions_fig = px.bar(top_functions, x=top_functions.values, y=top_functions.index, orientation='h', title='Top 10 Functions Today', labels={'x': 'Count', 'y': 'Function'})

    top_employees = df['team_member'].value_counts().nlargest(10).sort_values(ascending=True)
    top_employees_fig = px.bar(top_employees, x=top_employees.values, y=top_employees.index, orientation='h', title='Top 10 Employees Today', labels={'x': 'Count', 'y': 'Employee'})

    function_dist = df['function'].value_counts()
    function_dist_fig = px.pie(function_dist, values=function_dist.values, names=function_dist.index, title='Functions Distribution Today', hole=0.4)

    # --- 3. Save Figures to Image Bytes ---
    img_top_funcs = top_functions_fig.to_image(format="png", width=600, height=400)
    img_top_emps = top_employees_fig.to_image(format="png", width=600, height=400)
    img_func_dist = function_dist_fig.to_image(format="png", width=800, height=500)

    # --- 4. Create and Format Excel File in Memory ---
    output = io.BytesIO()
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Daily Analytics"

    # Title
    ws['A1'] = f"Daily Analytics Report for {today.strftime('%Y-%m-%d')}"
    ws['A1'].font = Font(size=20, bold=True)
    ws.merge_cells('A1:F1')

    # KPIs
    ws['A3'] = "Key Performance Indicators"; ws['A3'].font = Font(size=14, bold=True, underline="single")
    ws['A5'] = "Total Logs Today:"; ws['A5'].font = Font(bold=True)
    ws['B5'] = total_logs
    ws['A6'] = "Completion Rate:"; ws['A6'].font = Font(bold=True)
    ws['B6'] = f"{completion_rate:.2f}%"
    ws['A7'] = "Top Employee Today:"; ws['A7'].font = Font(bold=True)
    ws['B7'] = top_employee

    # Graphical Analysis
    ws['A9'] = "Graphical Analysis"; ws['A9'].font = Font(size=14, bold=True, underline="single")
    ws['A10'] = "Top 10 Functions by Log Count"; ws['A10'].font = Font(bold=True)
    ws.add_image(Image(io.BytesIO(img_top_funcs)), 'A11')
    ws['J10'] = "Top 10 Employees by Log Count"; ws['J10'].font = Font(bold=True)
    ws.add_image(Image(io.BytesIO(img_top_emps)), 'J11')
    ws['A34'] = "Overall Function Distribution"; ws['A34'].font = Font(bold=True)
    ws.add_image(Image(io.BytesIO(img_func_dist)), 'A35')

    workbook.save(output)
    output.seek(0)

    # --- 5. Serve the file ---
    filename = f"Daily_Analytics_{today.strftime('%Y-%m-%d')}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@app.route('/')
def landing():
    return render_template('login.html')

@app.route('/summary')
@login_required
def summary():
    if session.get('role') == 'admin':
        return redirect(url_for('admin_summary'))
    elif session.get('role') == 'employee':
        return redirect(url_for('employee_summary'))
    else:
        flash('You do not have permission to view a summary.', 'danger')
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # 1. IP-level Rate Limiting
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()
        if is_rate_limited(client_ip, limit=10, period=60):
            flash('Too many login attempts. Please try again in a minute.', 'danger')
            try:
                add_system_alert(f"Rate limit exceeded on login endpoint from IP: {client_ip}")
            except Exception:
                pass
            return redirect(url_for('login'))
        record_attempt(client_ip)

        username = request.form.get('username', '').strip()
        password = request.form.get('password')

        # Case-insensitive username lookup
        user = User.query.filter(func.lower(User.username) == username.lower()).first()

        if user:
            # 2. Account Lockout Check
            if user.lockout_until and user.lockout_until > datetime.now():
                lockout_left = int((user.lockout_until - datetime.now()).total_seconds() / 60) + 1
                flash(f"Account is temporarily locked. Try again in {lockout_left} minutes.", 'danger')
                try:
                    add_system_alert(f"Blocked login attempt for locked account: {user.username}")
                except Exception:
                    pass
                return redirect(url_for('login'))

            # 3. Password Verification
            if check_password_hash(user.password, password):
                # Reset failed login count and lockout state on success
                user.failed_login_attempts = 0
                user.lockout_until = None
                db.session.commit()

                # Enable session expiry timeout (starts permanent timer of 30 mins)
                session.permanent = True
                session['user'] = user.username
                session['role'] = user.role
                
                try:
                    if user.role == 'employee':
                        add_system_alert(f"Employee {user.username} logged in.")
                    elif user.role == 'admin':
                        add_system_alert(f"Admin {user.username} logged in.")
                except Exception:
                    pass
                
                if user.role == 'admin':
                    return redirect(url_for('admin_dashboard'))
                else:
                    return redirect(url_for('employee_dashboard'))
            else:
                # Increment failed login count
                user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
                if user.failed_login_attempts >= 5:
                    user.lockout_until = datetime.now() + timedelta(minutes=15)
                    flash('Invalid username or password. Your account has been locked for 15 minutes.', 'danger')
                    try:
                        add_system_alert(f"Account locked due to multiple failed login attempts: {user.username}")
                    except Exception:
                        pass
                else:
                    attempts_left = 5 - user.failed_login_attempts
                    flash(f'Invalid username or password. {attempts_left} attempts remaining.', 'danger')
                
                db.session.commit()
                try:
                    add_system_alert(f"Failed login attempt for user: {user.username}")
                except Exception:
                    pass
                return redirect(url_for('login'))
        else:
            # Prevent username timing attack using dummy verification
            check_password_hash(generate_password_hash('dummy_pass'), password)
            flash('Invalid username or password.', 'danger')
            try:
                add_system_alert(f"Failed login attempt for non-existent user: {username}")
            except Exception:
                pass
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'user' in session:
        # Log the logout for both admins and employees
        username = session.get('user')
        role = session.get('role')
        try:
            add_system_alert(f"{role.capitalize()} {username} logged out.")
        except Exception:
            pass
        session.clear()
    return redirect(url_for('login'))

@app.route('/employee/update', methods=['GET', 'POST'])
@login_required
def employee_update():
    current_user = User.query.filter_by(username=session.get('user')).first()
    user_department = current_user.department if current_user else ""

    if request.method == 'POST':
        try:
            date_str = request.form.get('date')
            log_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else None

            team_member = session.get('user', 'Guest')
            file_number = request.form.get('file_number')
            status = request.form.get('status')

            # Issue #4 fix: Check for ANY open 'In Progress' row (not just the latest),
            # so orphaned rows (from exceptions mid-session) also block new entries.
            if status == 'In Progress':
                open_log = Log.query.filter_by(
                    team_member=team_member,
                    status='In Progress'
                ).first()
                if open_log:
                    flash(
                        f"You cannot start a new file. Task for file "
                        f"'{open_log.file_number}' is still 'In Progress'.",
                        'danger'
                    )
                    return redirect(url_for('employee_update'))

            # Handle tasks without a file number, or when status is 'In Progress'.
            # These always create a new log entry.
            if not file_number or status == 'In Progress':
                month_str = log_date.strftime('%b-%y') if log_date else None
                new_log = Log(
                    team_member=team_member,
                    function=request.form.get('function'),
                    date=log_date,
                    file_number=file_number,
                    status=status,
                    tier1_escalation_reason=request.form.get('tier1_escalation'),
                    im_escalation_reason=request.form.get('im_escalation'),
                    department=user_department,
                    comments=request.form.get('comments'),
                    count=1,  # Issue #3: Integer, not string
                    bucket=request.form.get('function'),  # Same as function name
                    time=request.form.get('time'),
                    production_task=request.form.get('production_task'),
                    month=month_str
                )
                db.session.add(new_log)
                # Issue #4: Use a savepoint so a unique-index violation from a
                # concurrent duplicate request is caught cleanly without corrupting
                # the outer session transaction.
                try:
                    sp = db.session.begin_nested()
                    db.session.flush()
                    sp.commit()
                    db.session.commit()
                    flash('Work log added successfully!', 'success')
                except Exception:
                    db.session.rollback()
                    flash(
                        "A duplicate 'In Progress' entry was detected and blocked. "
                        "Please refresh and try again.",
                        'danger'
                    )
                    return redirect(url_for('employee_update'))

            # Handle status updates for existing 'In Progress' files
            else:  # file_number exists and status is not 'In Progress'
                log_to_update = Log.query.filter_by(
                    team_member=team_member,
                    file_number=file_number,
                    status='In Progress'
                ).first()

                if log_to_update:
                    month_str = log_date.strftime('%b-%y') if log_date else None
                    log_to_update.status = status
                    log_to_update.date = log_date
                    log_to_update.tier1_escalation_reason = request.form.get('tier1_escalation')
                    log_to_update.im_escalation_reason = request.form.get('im_escalation')
                    log_to_update.comments = request.form.get('comments')
                    log_to_update.count = 1  # Issue #3: Integer, not string
                    log_to_update.bucket = request.form.get('function')
                    log_to_update.production_task = request.form.get('production_task')
                    log_to_update.month = month_str
                    db.session.commit()
                    flash(f"Work log for file '{file_number}' updated to '{status}'.", 'success')
                else:
                    flash(
                        f"Error: You must first log file '{file_number}' with "
                        f"'In Progress' status before setting it to '{status}'.",
                        'danger'
                    )

        except Exception as e:
            db.session.rollback()
            flash(f'Error saving data: {str(e)}', 'danger')
        return redirect(url_for('employee_update'))

    # Fetch logs for the current user
    logs = Log.query.filter_by(
        team_member=session.get('user', 'Guest')
    ).order_by(
        Log.id.desc()
    ).all()

    functions = [f.name for f in Function.query.order_by(Function.name).all()]

    return render_template('employee/update_work.html', employee_name=session.get('user', 'Guest'), logs=logs, user_department=user_department, functions=functions)

@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    return render_template('employee/dashboard.html')

@app.route('/employee/summary')
@login_required
def employee_summary():
    # This page is for employees only.
    if session.get('role') != 'employee':
        flash('Access denied.', 'danger')
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('login'))

    logs = Log.query.filter_by(team_member=session.get('user')).all()

    all_functions = [f.name for f in Function.query.order_by(Function.name).all()]
    summary_counts = {func: 0 for func in all_functions}
    for log in logs:
        function = log.function
        if function in summary_counts:
            summary_counts[function] += 1
    
    functions = sorted(summary_counts.keys())

    return render_template('employee/summary.html', 
                           summary_counts=summary_counts, 
                           functions=functions, 
                           employee_name=session.get('user'))

@app.route('/admin/summary')
@admin_required
def admin_summary():
    # Using a more efficient query
    summary_counts_query = db.session.query(Log.function, func.count(Log.function)).group_by(Log.function).all()
    
    all_functions = [f.name for f in Function.query.order_by(Function.name).all()]
    summary_counts = {func: 0 for func in all_functions}
    for function, count in summary_counts_query:
        if function in summary_counts:
            summary_counts[function] = count
    
    functions = sorted(summary_counts.keys())

    return render_template('admin/summary.html', summary_counts=summary_counts, functions=functions)

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    alerts = Alert.query.order_by(Alert.timestamp.desc()).all()
    return render_template('admin/dashboard.html', alerts=alerts)

@app.route('/admin/functions', methods=['GET', 'POST'])
@admin_required
def manage_functions():
    if request.method == 'POST':
        function_name = request.form.get('name', '').strip()
        if not function_name:
            flash('Function name cannot be empty.', 'danger')
        elif Function.query.filter(func.lower(Function.name) == function_name.lower()).first():
            flash(f'Function "{function_name}" already exists.', 'danger')
        else:
            new_function = Function(name=function_name)
            db.session.add(new_function)
            db.session.commit()
            add_system_alert(f"Admin {session.get('user')} created function: {function_name}")
            flash(f'Function "{function_name}" created successfully.', 'success')
        return redirect(url_for('manage_functions'))

    functions = Function.query.order_by(Function.name).all()
    
    # Get department names from the Department table
    departments_query = Department.query.order_by(Department.dept_name).all()
    departments = [d.dept_name for d in departments_query if d.dept_name and d.dept_name.strip()]
    return render_template('admin/functions.html', functions=functions, departments=departments)

@app.route('/admin/functions/edit/<int:id>', methods=['POST'])
@admin_required
def edit_function(id):
    function_to_edit = Function.query.get_or_404(id)
    new_name = request.form.get('name', '').strip()

    if not new_name:
        flash('Function name cannot be empty.', 'danger')
    else:
        existing_function = Function.query.filter(func.lower(Function.name) == new_name.lower()).first()
        if existing_function and existing_function.id != id:
            flash(f'Function "{new_name}" already exists.', 'danger')
        else:
            old_name = function_to_edit.name
            Log.query.filter_by(function=old_name).update({'function': new_name})
            function_to_edit.name = new_name
            db.session.commit()
            add_system_alert(f"Admin {session.get('user')} edited function from '{old_name}' to '{new_name}'")
            flash(f'Function updated from "{old_name}" to "{new_name}".', 'success')
    return redirect(url_for('manage_functions'))

@app.route('/admin/functions/delete/<int:id>', methods=['POST'])
@admin_required
def delete_function(id):
    function_to_delete = Function.query.get_or_404(id)
    if Log.query.filter_by(function=function_to_delete.name).first():
        flash(f'Cannot delete function "{function_to_delete.name}" because it is in use in logs. Please edit it instead.', 'danger')
    else:
        func_name = function_to_delete.name
        db.session.delete(function_to_delete)
        db.session.commit()
        add_system_alert(f"Admin {session.get('user')} deleted function: {func_name}")
        flash(f'Function "{func_name}" has been deleted.', 'success')
    return redirect(url_for('manage_functions'))

@app.route('/admin/create_employee', methods=['GET', 'POST'])
@admin_required
def create_employee():
    if request.method == 'POST':
        username = request.form.get('team_member', '').strip()
        employee_id = request.form.get('employee_id', '').strip()
        department = request.form.get('department')
        role = request.form.get('role')
        shift = request.form.get('shift')
        location = request.form.get('location')
        password = request.form.get('password')

        if not all([username, employee_id, department, role, shift, location, password]):
            flash('All fields are required.', 'danger')
            return redirect(url_for('create_employee'))

        is_valid, complexity_err = validate_password_complexity(password)
        if not is_valid:
            flash(complexity_err, 'danger')
            return redirect(url_for('create_employee'))

        existing_user = User.query.filter(
            (func.lower(User.username) == username.lower()) | (User.employee_id == employee_id)
        ).first()
        if existing_user:
            flash(f'Employee with name "{username}" or ID "{employee_id}" already exists.', 'danger')
            return redirect(url_for('create_employee'))

        hashed_password = generate_password_hash(password)
        
        new_user = User(
            username=username,
            employee_id=employee_id,
            department=department,
            role=role,
            shift=shift,
            location=location,
            password=hashed_password,
            created_at=datetime.now()
        )
        db.session.add(new_user)
        db.session.commit()
        add_system_alert(f"Admin {session.get('user')} created employee: {username} ({role})")
        flash(f'Employee "{username}" created successfully!', 'success')
        return redirect(url_for('view_employees'))

    departments = [d.dept_name for d in Department.query.order_by(Department.dept_name).all()]

    return render_template('admin/create_employee.html', departments=departments)

@app.route('/admin/view_employees')
@admin_required
def view_employees():
    users = User.query.order_by(User.username).all()

    # Get the most recent log date for each user in a single query
    subquery = db.session.query(
        Log.team_member,
        func.max(Log.date).label('last_log_date')
    ).group_by(Log.team_member).subquery()

    last_log_dates = {row.team_member: row.last_log_date for row in db.session.query(subquery).all()}

    employees = []
    seven_days_ago = datetime.now().date() - timedelta(days=7)

    for user in users:
        last_date = last_log_dates.get(user.username)
        status = 'Inactive'
        
        if last_date and last_date >= seven_days_ago:
            status = 'Active'

        employees.append({
            'Employee ID': user.employee_id or 'N/A',
            'Team Member': user.username,
            'Department': user.department,
            'Shift': user.shift,
            'Location': user.location,
            'Status': status,
            'Last_Login': last_date.strftime('%Y-%m-%d') if last_date else 'N/A'
        })
    return render_template('admin/view_employees.html',
                           employees=employees, users=users)


# ─── Issue #5: Employee Management Routes ───────────────────────────────────

@app.route('/admin/employees/<int:user_id>/edit', methods=['POST'])
@admin_required
def edit_employee(user_id):
    """Edit an employee's profile fields (username, department, shift, location, role)."""
    user = User.query.get_or_404(user_id)

    # Prevent editing the only admin
    if user.role == 'admin' and session.get('user') != user.username:
        flash('Cannot edit another admin account.', 'danger')
        return redirect(url_for('view_employees'))

    new_username = request.form.get('username', '').strip()
    new_department = request.form.get('department', '').strip()
    new_shift = request.form.get('shift', '').strip()
    new_location = request.form.get('location', '').strip()
    new_role = request.form.get('role', '').strip()

    if not new_username:
        flash('Username cannot be empty.', 'danger')
        return redirect(url_for('view_employees'))

    # Check uniqueness only if username changed
    if new_username.lower() != user.username.lower():
        conflict = User.query.filter(
            func.lower(User.username) == new_username.lower(),
            User.id != user_id
        ).first()
        if conflict:
            flash(f'Username "{new_username}" is already taken.', 'danger')
            return redirect(url_for('view_employees'))

    old_username = user.username
    user.username = new_username
    user.department = new_department or user.department
    user.shift = new_shift or user.shift
    user.location = new_location or user.location
    if new_role in ('admin', 'employee'):
        user.role = new_role

    db.session.commit()
    add_system_alert(
        f"Admin {session.get('user')} edited employee '{old_username}' "
        f"-> username='{new_username}', dept='{user.department}', "
        f"shift='{user.shift}', role='{user.role}'"
    )
    flash(f'Employee "{old_username}" updated successfully.', 'success')
    return redirect(url_for('view_employees'))


@app.route('/admin/employees/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_employee(user_id):
    """Delete an employee account. Admins cannot delete themselves."""
    user = User.query.get_or_404(user_id)

    if user.username == session.get('user'):
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('view_employees'))

    if user.role == 'admin':
        flash('Cannot delete an admin account.', 'danger')
        return redirect(url_for('view_employees'))

    username = user.username
    db.session.delete(user)
    db.session.commit()
    add_system_alert(f"Admin {session.get('user')} deleted employee: {username}")
    flash(f'Employee "{username}" has been deleted.', 'success')
    return redirect(url_for('view_employees'))


@app.route('/admin/employees/<int:user_id>/reset_password', methods=['POST'])
@admin_required
def reset_employee_password(user_id):
    """Admin resets an employee's password with full complexity validation."""
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password', '')

    is_valid, err_msg = validate_password_complexity(new_password)
    if not is_valid:
        flash(f'Password reset failed: {err_msg}', 'danger')
        return redirect(url_for('view_employees'))

    user.password = generate_password_hash(new_password)
    user.failed_login_attempts = 0
    user.lockout_until = None
    db.session.commit()
    add_system_alert(
        f"Admin {session.get('user')} reset password for employee: {user.username}"
    )
    flash(f'Password for "{user.username}" has been reset successfully.', 'success')
    return redirect(url_for('view_employees'))

# ────────────────────────────────────────────────────────────────────────────


@app.route('/admin/production_report')
@admin_required
def production_report():
    from calendar import monthrange

    # --- Date-wise Production Data ---
    today = datetime.utcnow()
    
    # Determine the default year and month based on the most recent log in the database
    # this handles cases where the current month has no log data (e.g. July 2026)
    default_year = today.year
    default_month = today.month
    try:
        latest_log = Log.query.order_by(Log.date.desc()).first()
        if latest_log and latest_log.date:
            default_year = latest_log.date.year
            default_month = latest_log.date.month
    except Exception:
        pass

    # Allow overriding month/year via query params for filtering
    try:
        year_param = request.args.get('year')
        month_param = request.args.get('month')
        year = int(year_param) if year_param is not None else default_year
        month = int(month_param) if month_param is not None else default_month
        # Basic validation
        if not (1 <= month <= 12):
            month = default_month
        if not (2020 <= year <= today.year):
            year = default_year
    except (ValueError, TypeError):
        year = default_year
        month = default_month

    # Get number of days in the selected month
    num_days = monthrange(year, month)[1]
    days_in_month = list(range(1, num_days + 1))

    # Get all functions from Function master table
    functions = Function.query.order_by(Function.name).all()
    func_names = [f.name for f in functions if f.name]

    # Dynamically supplement with any unique functions from logs to ensure no data is missed
    try:
        log_func_query = db.session.query(Log.function).distinct().all()
        log_funcs = [f[0] for f in log_func_query if f[0]]
        for lf in log_funcs:
            if lf not in func_names:
                func_names.append(lf)
    except Exception:
        pass
    func_names = sorted(list(set(func_names)))

    # Query for logs in the selected month (Functions)
    daily_counts_query = db.session.query(
        Log.function,
        extract('day', Log.date).label('day'),
        func.count(Log.id).label('count')
    ).filter(
        extract('year', Log.date) == year,
        extract('month', Log.date) == month,
        Log.function.in_(func_names)
    ).group_by(Log.function, extract('day', Log.date)).all()

    # Process data into a pivot-table like structure
    data = {f_name: {day: 0 for day in days_in_month} for f_name in func_names}
    for function_name, day, count in daily_counts_query:
        if function_name in data and day in data[function_name]:
            data[function_name][int(day)] = count

    # Prepare final list for template, including totals
    production_by_date = []
    for func_name, daily_counts in data.items():
        production_by_date.append({
            'function': func_name,
            'days': daily_counts,
            'total': sum(daily_counts.values())
        })

    month_name = datetime(year, month, 1).strftime('%B')
    years = list(range(today.year, 2019, -1))
    months = {i: datetime(2000, i, 1).strftime('%B') for i in range(1, 13)}
    
    return render_template('admin/production_report.html', production_by_date=production_by_date, days_in_month=days_in_month, month_name=month_name, selected_year=year, selected_month=month, years=years, months=months)

@app.route('/admin/production_by_department')
@admin_required
def production_by_department():
    # Get all departments from the master Department table
    departments = Department.query.order_by(Department.dept_name).all()
    existing_dept_names = {d.dept_name for d in departments if d.dept_name}

    # Get total logs per department in a single query
    dept_logs_query = db.session.query(
        Log.department,
        func.count(Log.id)
    ).group_by(Log.department).all()
    dept_logs_map = dict(dept_logs_query)

    # Efficiently get the top function for each department using a window function
    # This avoids making a query for each department inside a loop (N+1 problem)
    
    # Subquery to count functions per department
    log_counts_subquery = db.session.query(
        Log.department,
        Log.function,
        func.count(Log.id).label('function_count')
    ).filter(Log.department.isnot(None)).group_by(Log.department, Log.function).subquery()

    # Window function to rank functions within each department
    ranked_logs_subquery = db.session.query(
        log_counts_subquery.c.department,
        log_counts_subquery.c.function,
        func.row_number().over(
            partition_by=log_counts_subquery.c.department,
            order_by=log_counts_subquery.c.function_count.desc()
        ).label('rn')
    ).subquery()

    # Select only the top-ranked function (rn=1) for each department
    top_functions_query = db.session.query(
        ranked_logs_subquery.c.department,
        ranked_logs_subquery.c.function
    ).filter(ranked_logs_subquery.c.rn == 1).all()
    top_functions_map = dict(top_functions_query)

    # Combine departments from both master list and logs to ensure complete stats
    dept_stats_pre = []
    processed_depts = set()
    
    for dept in departments:
        dept_name = dept.dept_name
        dept_stats_pre.append({
            'department': dept_name,
            'total_logs': dept_logs_map.get(dept_name, 0),
            'top_function': None
        })
        processed_depts.add(dept_name)

    for dept_name in dept_logs_map.keys():
        if dept_name and dept_name not in processed_depts:
            dept_stats_pre.append({
                'department': dept_name,
                'total_logs': dept_logs_map.get(dept_name, 0),
                'top_function': None
            })
            processed_depts.add(dept_name)

    department_stats = []
    for stat in dept_stats_pre:
        dept_name = stat['department']
        stat['top_function'] = top_functions_map.get(dept_name)
        department_stats.append(stat)

    return render_template('admin/production_by_department.html', department_stats=department_stats)

@app.route('/admin/team_member_performance')
@admin_required
def team_member_performance():
    # Get all employee users
    users = User.query.filter(User.role == 'employee').order_by(User.username).all()
    
    # Get performance stats in one subquery
    performance_query = db.session.query(
        Log.team_member,
        func.count(Log.id).label('total_logs'),
        func.count(func.distinct(Log.date)).label('active_days'),
        func.max(Log.date).label('last_log_date')
    ).group_by(Log.team_member).subquery()

    # Create a dictionary for easy lookup
    performance_stats = {
        row.team_member: {
            'total_logs': row.total_logs,
            'active_days': row.active_days,
            'last_log_date': row.last_log_date,
            'avg_per_day': (row.total_logs / row.active_days) if row.active_days > 0 else 0
        }
        for row in db.session.query(performance_query).all()
    }

    performance_data = []
    for user in users:
        stats = performance_stats.get(user.username, {'total_logs': 0, 'avg_per_day': 0, 'last_log_date': None})
        performance_data.append({'username': user.username, 'department': user.department, **stats})

    return render_template('admin/team_member_performance.html', performance_data=performance_data)

@app.route('/admin/tracker')
@admin_required
def track_employee():
    users = User.query.order_by(User.username).all()
    employees = [user.username for user in users]

    selected_employee = request.args.get('employee')
    page = request.args.get('page', 1, type=int)
    
    query = Log.query
    if selected_employee:
        query = query.filter_by(team_member=selected_employee)
        
    # Add pagination to the query
    pagination = query.order_by(Log.date.desc(), Log.id.desc()).paginate(page=page, per_page=100, error_out=False)
    logs_to_display = pagination.items

    # Calculate Statistics if an employee is selected, using SQL aggregations
    stats = {
        'avg_per_day': 0,
        'top_function': 'N/A',
        'function_breakdown': {}
    }

    if selected_employee:
        # 1. Function Breakdown and Top Function (SQL)
        function_breakdown_query = db.session.query(
            Log.function, 
            func.count(Log.id)
        ).filter(
            Log.team_member == selected_employee
        ).group_by(
            Log.function
        ).order_by(
            func.count(Log.id).desc()
        ).all()

        if function_breakdown_query:
            stats['function_breakdown'] = dict(function_breakdown_query)
            stats['top_function'] = function_breakdown_query[0][0]

        # 2. Average files per day (SQL)
        daily_counts_subquery = db.session.query(
            func.count(Log.id).label('daily_count')
        ).filter(Log.team_member == selected_employee).group_by(Log.date).subquery()

        avg_per_day_query = db.session.query(func.avg(daily_counts_subquery.c.daily_count)).scalar()
        
        if avg_per_day_query is not None:
            stats['avg_per_day'] = round(float(avg_per_day_query), 2)

    return render_template('admin/track_employee.html', employees=employees, logs=logs_to_display, selected_employee=selected_employee, stats=stats, pagination=pagination)

@app.route('/admin/tracker/export')
@admin_required
def export_tracker_data():
    selected_employee = request.args.get('employee')
    if not selected_employee:
        flash('Please select an employee to export data.', 'warning')
        return redirect(url_for('track_employee'))

    # Fetch logs for the selected employee, ordered by date
    logs = Log.query.filter_by(team_member=selected_employee).order_by(Log.date.asc()).all()

    if not logs:
        flash(f'No logs found for {selected_employee} to export.', 'info')
        return redirect(url_for('track_employee', employee=selected_employee))

    # Create a pandas DataFrame from the log data
    df = pd.DataFrame(
        [
            {
                "Date": log.date,
                "Function": log.function,
                "File Number": log.file_number,
                "Status": log.status,
                "Department": log.department,
                "Comments": log.comments,
                "Count": log.count,
                "Bucket": log.bucket,
                "Production Task": log.production_task,
                "Month": log.month,
            }
            for log in logs
        ]
    )

    # --- Create Excel Data in Memory ---
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Detailed Logs
        df.to_excel(writer, sheet_name='Detailed Logs', index=False)

        # Sheet 2: Date-wise file count (Daily Summary)
        if 'Date' in df.columns:
            df_date_summary = df.groupby('Date').size().reset_index(name='Files Count')
            df_date_summary.to_excel(writer, sheet_name='Daily Summary', index=False)

        # Sheet 3: Function distribution
        if 'Function' in df.columns:
            df_func_dist = df['Function'].value_counts().reset_index()
            df_func_dist.columns = ['Function', 'Count']
            df_func_dist.to_excel(writer, sheet_name='Function Distribution', index=False)

    output.seek(0)

    # --- Serve the file for download ---
    filename = f"{selected_employee}_Tracker_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@app.route('/logo.png')
def serve_logo():
    return send_from_directory(os.path.join(BASE_DIR, 'templates'), 'logo.png')

@app.route('/favicon.ico')
def favicon():
    # Handle browser request for favicon to prevent 404 error
    return '', 204

@app.route('/chart-data')
@admin_required
def chart_data():
    try:
        # Query logs and aggregate by date and function
        logs_by_date_func = db.session.query(
            Log.date,
            Log.function,
            func.count(Log.id)
        ).group_by(Log.date, Log.function).all()

        all_functions = [f.name for f in Function.query.order_by(Function.name).all()]
        columns = all_functions + ["Total Hours"] # Total Hours seems unused, keeping for compatibility

        # Aggregate data by Date
        aggregated_data = {}
        for log_date, function, count in logs_by_date_func:
            if not log_date:
                continue
            
            date_str = log_date.strftime('%Y-%m-%d')
                
            if date_str not in aggregated_data:
                # Initialize row with 0s
                row = {col: 0 for col in columns}
                row['Date'] = date_str
                aggregated_data[date_str] = row
            
            # Increment count if the function matches a column
            if function in aggregated_data[date_str]:
                aggregated_data[date_str][function] += count

        # Convert dict to list
        return jsonify(list(aggregated_data.values()))
    except Exception as e:
        print(f"Error generating chart data: {e}")
        return jsonify({'error': str(e)})

@app.cli.command("init-db")
def init_db_command():
    """Creates the database tables and a default admin user."""
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        hashed_password = generate_password_hash('admin')
        admin_user = User(username='admin', password=hashed_password, role='admin', department='System')
        db.session.add(admin_user)
        db.session.commit()
        print("Database initialized and admin user created.")
    else:
        print("Database already initialized.")

    if not Function.query.first():
        print("Populating functions table...")
        default_functions = [
            "VI 3D Scan Pro", "VI 3D Desktop Pro", "Full Review", "Full Revision",
            "Short Review", "Short Revision", "VI Second Review",
            "Digital Operations - Sourcing", "Full Reports", "QCF (Underwriter Queue)",
            "Full Review (CI Abridged)", "CMP Client Import", "Text Followup", "ACR",
            "DNU Checklist Update", "PDC Compliance", "Meetings/Training"
        ]
        for func_name in default_functions:
            db.session.add(Function(name=func_name))
        db.session.commit()
        print("Functions table populated.")

    if not Department.query.first():
        print("Populating department table...")
        default_departments = [
            'Alternative Products',
            'Assigning',
            'Bluebird QC',
            'Client Services',
            'Digital',
            'Management',
            'Quality Control',
            'Staff Direct',
            'Training',
            'Vendor Relations'
        ]
        for dept_name in default_departments:
            db.session.add(Department(dept_name=dept_name))
        db.session.commit()
        print("Department table populated.")

@app.cli.command("import-data")
@click.option('--yes', is_flag=True, default=False, help='Confirm DROP of all log data and re-import. Required in full-replace mode.')
@click.option('--file', default=None, help='Name of the Excel file in the project folder to import (e.g. "Production & Performance Report till Aug 5th.xlsx").')
@click.option('--append', is_flag=True, default=False, help='Append new data without dropping existing logs. Safe for adding a new month on top of existing data.')
def import_data_command(yes, file, append):
    """Imports users and log data from a production report Excel file.

    Full-replace mode (drops existing logs):
      flask import-data --file "MyFile.xlsx" --yes

    Append mode (keeps existing logs, adds new rows):
      flask import-data --file "MyFile.xlsx" --append
    """
    if not append and not yes:
        print("\n[WARNING] This command will DROP the entire Log table and re-import all data.")
        print("   All existing log rows will be PERMANENTLY DELETED.")
        print("   To confirm a full replace: flask import-data --file \"MyFile.xlsx\" --yes")
        print("   To append without deleting: flask import-data --file \"MyFile.xlsx\" --append\n")
        return

    new_report_file = os.path.join(BASE_DIR, file) if file else os.path.join(BASE_DIR, 'Production & Performance Report till Aug 5th.xlsx')

    if not new_report_file.lower().endswith('.xlsx'):
        print(f"Error: File must be a .xlsx spreadsheet.")
        return
    if not os.path.exists(new_report_file):
        print(f"Error: Report file not found: {new_report_file}")
        return

    mode_label = "[APPEND MODE]" if append else "[FULL REPLACE MODE]"
    print(f"\n--- {mode_label} Starting import from: {os.path.basename(new_report_file)} ---")

    try:
        xls = pd.ExcelFile(new_report_file)
        if 'Team Member Performance' not in xls.sheet_names or 'Raw Data' not in xls.sheet_names:
            print(f"Error: Required sheets not found. Available: {xls.sheet_names}")
            return

        print("Reading Excel sheets (this may take a moment for large files)...")
        df_performance = pd.read_excel(xls, sheet_name='Team Member Performance')
        df_raw = pd.read_excel(xls, sheet_name='Raw Data')
        df_performance.columns = df_performance.columns.str.strip()
        df_raw.columns = df_raw.columns.str.strip()
        print(f"Read {len(df_raw)} rows from Raw Data sheet.")

        # --- Sync Departments ---
        unique_departments = df_raw['Department'].dropna().unique()
        existing_depts = {d.dept_name for d in Department.query.all()}
        depts_added = 0
        for dept_name in unique_departments:
            if dept_name not in existing_depts:
                db.session.add(Department(dept_name=dept_name))
                depts_added += 1
        db.session.commit()
        print(f"Departments synced ({depts_added} new added).")

        # --- Department mapping for user import ---
        department_map = {}
        if 'Team Member (First Last)' in df_raw.columns and 'Department' in df_raw.columns:
            df_dept = df_raw[['Team Member (First Last)', 'Department']].dropna()
            df_dept['Team Member (First Last)'] = df_dept['Team Member (First Last)'].str.strip()
            department_map = df_dept.drop_duplicates('Team Member (First Last)', keep='first').set_index('Team Member (First Last)')['Department'].to_dict()

        # --- Import Users ---
        df_baroda = df_performance[df_performance['Branch'] == 'Baroda'].copy() if 'Branch' in df_performance.columns else pd.DataFrame()
        print(f"Found {len(df_baroda)} users for 'Baroda' branch.")

        if append:
            # Append mode: only add new users, keep existing ones untouched
            existing_usernames = {u.username.lower() for u in User.query.all()}
            users_added = 0
            for _, row in df_baroda.iterrows():
                username = str(row.get('Team Member (First Last)', '')).strip()
                employee_id = str(row.get('Employee ID', '')).strip()
                if not username or username.lower() in ['nan', ''] or not employee_id:
                    continue
                if username.lower() in existing_usernames:
                    continue
                db.session.add(User(
                    username=username, employee_id=employee_id,
                    password=generate_password_hash('password'), role='employee',
                    department=department_map.get(username),
                    shift=str(row.get('Shift', '')).strip(), location='Baroda',
                    created_at=datetime.now()
                ))
                users_added += 1
            db.session.commit()
            print(f"Users: {users_added} new added (existing users untouched).")
        else:
            # Full replace mode: delete all non-admin users and re-create
            num_deleted = User.query.filter(User.role != 'admin').delete()
            db.session.commit()
            print(f"Deleted {num_deleted} existing non-admin users.")
            users_added = 0
            for _, row in df_baroda.iterrows():
                username = str(row.get('Team Member (First Last)', '')).strip()
                employee_id = str(row.get('Employee ID', '')).strip()
                if not username or username.lower() in ['nan', ''] or not employee_id:
                    continue
                db.session.add(User(
                    username=username, employee_id=employee_id,
                    password=generate_password_hash('password'), role='employee',
                    department=department_map.get(username),
                    shift=str(row.get('Shift', '')).strip(), location='Baroda',
                    created_at=datetime.now()
                ))
                users_added += 1
            db.session.commit()
            print(f"Users: {users_added} imported.")

        # --- Import Logs ---
        print("\n--- Starting Log import ---")
        if not append:
            # Full replace: backup then DROP
            try:
                existing_logs = Log.query.all()
                if existing_logs:
                    backup_dir = os.path.join(BASE_DIR, 'backups')
                    os.makedirs(backup_dir, exist_ok=True)
                    backup_filename = f"log_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    backup_path = os.path.join(backup_dir, backup_filename)
                    df_backup = pd.DataFrame([{
                        'id': l.id, 'team_member': l.team_member, 'function': l.function,
                        'date': l.date, 'file_number': l.file_number, 'status': l.status,
                        'department': l.department, 'count': l.count, 'bucket': l.bucket,
                        'time': l.time, 'production_task': l.production_task, 'month': l.month,
                        'comments': l.comments,
                    } for l in existing_logs])
                    df_backup.to_excel(backup_path, index=False)
                    print(f"[OK] Backup saved: {backup_path} ({len(existing_logs)} rows)")
                else:
                    print("No existing logs to back up.")
            except Exception as backup_err:
                print(f"[WARNING] Backup failed: {backup_err}. Aborting to protect data.")
                return
            db.session.execute(text('DROP TABLE IF EXISTS log CASCADE;'))
            db.session.commit()
            db.create_all()
            print("Log table dropped and recreated.")
        else:
            existing_count = Log.query.count()
            file_months = list(df_raw['Month'].dropna().unique()) if 'Month' in df_raw.columns else []
            existing_months = [m[0] for m in db.session.query(Log.month).distinct().all() if m[0]]
            overlap = set(file_months) & set(existing_months)
            if overlap:
                print(f"[WARNING] Month(s) {list(overlap)} already in DB. Rows will be added on top.")
            print(f"Appending to {existing_count} existing log rows...")

        # Build and insert log rows in batches
        required_cols = ['Team Member (First Last)', 'Date (mm/dd/yy)', 'Function']
        if not all(c in df_raw.columns for c in required_cols):
            print(f"Error: Missing required columns. Found: {list(df_raw.columns)}")
            return

        df_raw_logs = df_raw.where(pd.notnull(df_raw), None)
        logs_to_add = []
        for _, row in df_raw_logs.iterrows():
            if not row['Team Member (First Last)'] or not row['Date (mm/dd/yy)']:
                continue
            try:
                log_date = pd.to_datetime(row['Date (mm/dd/yy)']).date()
            except (ValueError, TypeError):
                continue
            logs_to_add.append(Log(
                team_member=row.get('Team Member (First Last)'),
                function=row.get('Function'),
                date=log_date,
                file_number=str(row.get('File  Number')) if row.get('File  Number') else None,
                status=str(row.get('Status')) if pd.notna(row.get('Status')) else None,
                tier1_escalation_reason=str(row.get('Escalation Reason')) if pd.notna(row.get('Escalation Reason')) else None,
                im_escalation_reason=None,
                department=row.get('Department'),
                comments=None,
                count=int(row.get('Count')) if row.get('Count') and str(row.get('Count')).isdigit() else 1,
                bucket=row.get('Bucket'),
                time=str(row.get('Time')) if row.get('Time') else None,
                production_task=row.get('Production Task'),
                month=row.get('Month')
            ))

        if logs_to_add:
            batch_size = 500
            total_imported = 0
            for i in range(0, len(logs_to_add), batch_size):
                batch = logs_to_add[i:i + batch_size]
                db.session.bulk_save_objects(batch)
                db.session.commit()
                total_imported += len(batch)
                print(f"  Batch {i // batch_size + 1}: {total_imported}/{len(logs_to_add)} rows committed...")
            total_in_db = Log.query.count()
            print(f"\n--- Done: {total_imported} logs imported. Total in DB: {total_in_db} ---")
        else:
            print("No valid log entries found to import.")

    except Exception as e:
        db.session.rollback()
        print(f"An error occurred during data import: {e}")

@app.cli.command("count-rows")
def count_rows_command():
    """Counts and prints the number of rows in key tables."""
    try:
        log_count = Log.query.count()
        user_count = User.query.count()
        print("\n--- Database Row Counts ---")
        print(f"Log table:    {log_count} rows")
        print(f"User table:   {user_count} rows")
        print("---------------------------\n")
    except Exception as e:
        print(f"An error occurred while counting rows: {e}")

init_dashboard(app)
init_daily_dashboard(app)

@app.cli.command("audit-dependencies")
def audit_dependencies():
    """CLI tool for automated dependency vulnerability scanning."""
    import subprocess
    print("Running dependency vulnerability scan (pip-audit / safety)...")
    
    # Try running pip-audit
    try:
        res = subprocess.run(["pip-audit"], capture_output=True, text=True)
        if res.returncode == 0:
            print("pip-audit output (No vulnerabilities found):")
            print(res.stdout)
        else:
            print("Vulnerabilities detected by pip-audit:")
            print(res.stdout)
            print(res.stderr)
        return
    except FileNotFoundError:
        pass

    # Try running safety check
    try:
        res = subprocess.run(["safety", "check"], capture_output=True, text=True)
        if res.returncode == 0:
            print("safety check output (No vulnerabilities found):")
            print(res.stdout)
        else:
            print("Vulnerabilities detected by safety:")
            print(res.stdout)
            print(res.stderr)
        return
    except FileNotFoundError:
        pass

    print("Error: Neither 'pip-audit' nor 'safety' is installed.")
    print("To install, run: pip install pip-audit   OR   pip install safety")

if __name__ == '__main__':
    print(f"Template folder set to: {os.path.join(BASE_DIR, 'templates')}")
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    port = int(os.environ.get("PORT", 5000))

    # Use the dynamic port and bind to 0.0.0.0
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
